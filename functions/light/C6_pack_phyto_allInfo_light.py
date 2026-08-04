from __future__ import annotations

import argparse
import csv
import os
import re
import tempfile
import zipfile
from pathlib import Path

import h5py
import numpy as np
import scipy.io as sio
from openpyxl import load_workbook


_TAG_ALIASES = {"tag", "tag_number", "tagnumber", "event", "event_id", "eventid"}
_DELAY_ALIASES = {"delay", "delay_ps", "delaytime", "delay_time", "time_delay"}


def _load_h5_vectors(path, names):
    with h5py.File(path, "r") as h5:
        return [np.asarray(h5[n][()]).reshape(-1) for n in names]


def _normalise_header(value) -> str:
    text = str(value).strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _parse_tag(value, row_number: int) -> int:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"Missing tag at row {row_number}")

    if isinstance(value, (int, np.integer)):
        return int(value)
    if isinstance(value, (float, np.floating)):
        if not np.isfinite(value) or not float(value).is_integer():
            raise ValueError(f"Invalid non-integer tag {value!r} at row {row_number}")
        return int(value)

    text = str(value).strip()
    match = re.fullmatch(r"(?:tag-?)?(\d+)(?://)?", text, flags=re.IGNORECASE)
    if not match:
        raise ValueError(f"Invalid tag {value!r} at row {row_number}")
    return int(match.group(1))


def _parse_delay(value, row_number: int) -> float:
    try:
        delay = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid delay {value!r} at row {row_number}") from exc
    if not np.isfinite(delay):
        raise ValueError(f"Non-finite delay {value!r} at row {row_number}")
    return delay


def _header_indices(header) -> tuple[int, int]:
    normalised = [_normalise_header(value) for value in header]
    tag_index = next((i for i, name in enumerate(normalised) if name in _TAG_ALIASES), None)
    delay_index = next((i for i, name in enumerate(normalised) if name in _DELAY_ALIASES), None)
    if tag_index is None or delay_index is None:
        raise KeyError(
            "The tag-delay table must contain tag_number/tag and delay columns. "
            f"Detected columns: {normalised}"
        )
    return tag_index, delay_index


def _split_single_cell_row(value) -> list[str]:
    if value is None:
        return []
    return next(csv.reader([str(value)]))


def _load_tag_delay_xlsx(path: Path, sheet_name: str | None = None) -> tuple[np.ndarray, np.ndarray]:
    """Read either a normal multi-column worksheet or the supplied one-cell CSV-style worksheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"Worksheet {sheet_name!r} not found. Available: {workbook.sheetnames}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = worksheet.iter_rows(values_only=True)
        first_row = next((tuple(row) for row in rows if any(value is not None for value in row)), None)
        if first_row is None:
            raise ValueError(f"The tag-delay workbook is empty: {path}")

        single_cell_csv = len(first_row) == 1 and isinstance(first_row[0], str) and "," in first_row[0]
        header = _split_single_cell_row(first_row[0]) if single_cell_csv else list(first_row)
        tag_index, delay_index = _header_indices(header)

        tags: list[int] = []
        delays: list[float] = []
        for row_number, row in enumerate(rows, start=2):
            values = _split_single_cell_row(row[0]) if single_cell_csv else list(row)
            if not values or all(value is None or str(value).strip() == "" for value in values):
                continue
            required_index = max(tag_index, delay_index)
            if len(values) <= required_index:
                raise ValueError(f"Row {row_number} has too few columns: {values!r}")
            tags.append(_parse_tag(values[tag_index], row_number))
            delays.append(_parse_delay(values[delay_index], row_number))
    finally:
        workbook.close()

    return np.asarray(delays, dtype=np.float64), np.asarray(tags, dtype=np.int64)


def _load_tag_delay_text(path: Path) -> tuple[np.ndarray, np.ndarray]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t; ")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        header = next(reader, None)
        if header is None:
            raise ValueError(f"The tag-delay text file is empty: {path}")
        tag_index, delay_index = _header_indices(header)

        tags: list[int] = []
        delays: list[float] = []
        for row_number, row in enumerate(reader, start=2):
            if not row or all(not str(value).strip() for value in row):
                continue
            required_index = max(tag_index, delay_index)
            if len(row) <= required_index:
                raise ValueError(f"Row {row_number} has too few columns: {row!r}")
            tags.append(_parse_tag(row[tag_index], row_number))
            delays.append(_parse_delay(row[delay_index], row_number))

    return np.asarray(delays, dtype=np.float64), np.asarray(tags, dtype=np.int64)


def _load_tag_delay_mat(path: Path, data_name: str) -> tuple[np.ndarray, np.ndarray]:
    suffix = "1ps" if data_name in ("upto1ps", "1ps") else data_name
    try:
        mat = sio.loadmat(path)
        keys = {key: value for key, value in mat.items() if not key.startswith("__")}
        delay_key = f"delay_{suffix}" if f"delay_{suffix}" in keys else "delay"
        tag_key = f"tag_{suffix}" if f"tag_{suffix}" in keys else "tag"
        if delay_key not in keys or tag_key not in keys:
            raise KeyError(f"Could not find delay/tag for suffix {suffix}. Keys: {list(keys)}")
        return np.asarray(keys[delay_key]).squeeze(), np.asarray(keys[tag_key]).squeeze()
    except NotImplementedError:
        with h5py.File(path, "r") as h5:
            delay_key = f"delay_{suffix}" if f"delay_{suffix}" in h5 else "delay"
            tag_key = f"tag_{suffix}" if f"tag_{suffix}" in h5 else "tag"
            if delay_key not in h5 or tag_key not in h5:
                raise KeyError(f"Could not find delay/tag for suffix {suffix}. Keys: {list(h5.keys())}")
            return np.asarray(h5[delay_key][()]).squeeze(), np.asarray(h5[tag_key][()]).squeeze()


def _load_delay_file(path, data_name, sheet_name: str | None = None):
    """Load tag -> delay data from XLSX/CSV/TXT or the legacy MAT format.

    ZIP archives are accepted when they contain exactly one supported table file.
    """
    source = Path(path)
    suffix = source.suffix.lower()

    if suffix == ".zip":
        with zipfile.ZipFile(source) as archive:
            candidates = [
                name
                for name in archive.namelist()
                if not name.startswith("__MACOSX/")
                and Path(name).suffix.lower() in {".xlsx", ".xlsm", ".csv", ".txt", ".tsv", ".mat"}
            ]
            if len(candidates) != 1:
                raise ValueError(
                    f"Expected exactly one supported tag-delay file in {source}; found {candidates}"
                )
            with tempfile.TemporaryDirectory(prefix="tag_delay_") as temp_dir:
                extracted = Path(archive.extract(candidates[0], temp_dir))
                return _load_delay_file(extracted, data_name, sheet_name)

    if suffix in {".xlsx", ".xlsm"}:
        delays, tags = _load_tag_delay_xlsx(source, sheet_name)
    elif suffix in {".csv", ".txt", ".tsv"}:
        delays, tags = _load_tag_delay_text(source)
    elif suffix == ".mat":
        delays, tags = _load_tag_delay_mat(source, data_name)
    else:
        raise ValueError(
            f"Unsupported tag-delay format {suffix!r}. Use .xlsx, .csv, .txt, .tsv, .zip, or legacy .mat."
        )

    delays = np.asarray(delays, dtype=np.float64).reshape(-1)
    tags = np.asarray(tags).reshape(-1)
    if delays.size != tags.size:
        raise ValueError(f"Tag and delay lengths differ: {tags.size} tags vs {delays.size} delays")
    parsed_tags = np.asarray([_parse_tag(value, i + 2) for i, value in enumerate(tags)], dtype=np.int64)
    if not np.all(np.isfinite(delays)):
        raise ValueError("The delay file contains NaN or infinite values")
    return delays, parsed_tags


def _unique_map(keys, values, label: str):
    mapping = {}
    duplicate_same = 0
    for key, value in zip(keys, values):
        if key in mapping:
            old = mapping[key]
            if np.allclose(np.asarray(old, dtype=float), np.asarray(value, dtype=float), rtol=0.0, atol=1e-12):
                duplicate_same += 1
                continue
            raise ValueError(f"Conflicting duplicate {label} key {key}: {old!r} vs {value!r}")
        mapping[key] = value
    return mapping, duplicate_same


def pack_phyto_allInfo_light(
    stream_info_h5,
    params_h5,
    tag_delay_file,
    data_name,
    folder_path=".",
    delay_sheet: str | None = None,
):
    """Merge stream metadata, partialator parameters, and a tag-delay lookup table.

    Stream and params are joined by (runID, eventID). Delay is joined by eventID/tag.
    Output columns: runID, eventID, nBrg, DRL, delay, OSF, relB.
    """
    stream_run, stream_event, n_refl, drl = _load_h5_vectors(
        stream_info_h5, ["nRun", "nEvent", "nRefl", "DRL"]
    )
    params_run, params_event, osf, rel_b = _load_h5_vectors(
        params_h5, ["nRun", "nEvent", "OSF", "relB"]
    )

    stream_lengths = {len(np.atleast_1d(array)) for array in (stream_run, stream_event, n_refl, drl)}
    params_lengths = {len(np.atleast_1d(array)) for array in (params_run, params_event, osf, rel_b)}
    if len(stream_lengths) != 1:
        raise ValueError(f"C2 stream arrays have inconsistent lengths: {stream_lengths}")
    if len(params_lengths) != 1:
        raise ValueError(f"C1 params arrays have inconsistent lengths: {params_lengths}")

    params_keys = [(int(run), int(event)) for run, event in zip(params_run, params_event)]
    params_values = [(float(scale), float(b_factor)) for scale, b_factor in zip(osf, rel_b)]
    params_by_snapshot, duplicate_params = _unique_map(params_keys, params_values, "params snapshot")

    delay_all, tag_all = _load_delay_file(tag_delay_file, data_name, delay_sheet)
    delay_by_tag, duplicate_tags = _unique_map(
        [int(tag) for tag in tag_all], [float(delay) for delay in delay_all], "tag-delay"
    )

    rows = []
    missing_params = []
    missing_delay = []
    for run, event, refl, resolution in zip(stream_run, stream_event, n_refl, drl):
        snapshot_key = (int(run), int(event))
        if snapshot_key not in params_by_snapshot:
            missing_params.append(snapshot_key)
            continue
        event_int = snapshot_key[1]
        if event_int not in delay_by_tag:
            missing_delay.append(snapshot_key)
            continue
        scale, b_factor = params_by_snapshot[snapshot_key]
        rows.append(
            (
                snapshot_key[0],
                event_int,
                int(refl),
                float(resolution),
                delay_by_tag[event_int],
                scale,
                b_factor,
            )
        )

    if not rows:
        raise RuntimeError(
            "No common snapshots found across stream, params, and tag-delay data. "
            f"stream={len(np.atleast_1d(stream_run))}, params={len(params_by_snapshot)}, "
            f"tag-delay={len(delay_by_tag)}"
        )

    info = np.asarray(rows, dtype=np.float64)
    run_id = info[:, 0].astype(np.int64)
    event_id = info[:, 1].astype(np.int64)
    n_brg = info[:, 2].astype(np.int64)
    drl_out = info[:, 3]
    delay = info[:, 4]
    osf_out = info[:, 5]
    rel_b_out = info[:, 6]

    os.makedirs(folder_path, exist_ok=True)
    out_h5 = os.path.join(folder_path, f"merged_snapshotInfo_phyto_{data_name}_allInfo.hdf5")
    out_dat = os.path.join(folder_path, f"merged_snapshotInfo_phyto_{data_name}_allInfo.dat")

    with h5py.File(out_h5, "w") as h5:
        h5.create_dataset("runID", data=run_id)
        h5.create_dataset("eventID", data=event_id)
        h5.create_dataset("nBrg", data=n_brg)
        h5.create_dataset("DRL", data=drl_out)
        h5.create_dataset("delay", data=delay)
        h5.create_dataset("OSF", data=osf_out)
        h5.create_dataset("relB", data=rel_b_out)
        h5.create_dataset("notice", data=np.bytes_("parameters NOT sorted based on delay yet"))
        h5.attrs["tag_delay_source"] = str(Path(tag_delay_file).resolve())
        h5.attrs["matched_snapshots"] = len(rows)
        h5.attrs["missing_params_snapshots"] = len(missing_params)
        h5.attrs["missing_delay_snapshots"] = len(missing_delay)
        h5.attrs["duplicate_identical_params"] = duplicate_params
        h5.attrs["duplicate_identical_tags"] = duplicate_tags

    np.savetxt(
        out_dat,
        np.column_stack([run_id, event_id, n_brg, drl_out, delay, osf_out, rel_b_out]),
        fmt="%12d %16d %8d %10.4f %12.4f %12.6f %12.6f",
    )

    print(
        "C6 match summary: "
        f"stream={len(np.atleast_1d(stream_run))}, params={len(params_by_snapshot)}, "
        f"tag-delay={len(delay_by_tag)}, matched={len(rows)}, "
        f"missing_params={len(missing_params)}, missing_delay={len(missing_delay)}"
    )
    if missing_params:
        print(f"First missing params snapshots: {missing_params[:5]}")
    if missing_delay:
        print(f"First missing delay snapshots: {missing_delay[:5]}")
    print(f"Saved {out_h5}")
    print(f"Saved {out_dat}")
    return out_h5


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("stream_info_h5")
    parser.add_argument("params_h5")
    parser.add_argument("tag_delay_file")
    parser.add_argument("--data-name", required=True)
    parser.add_argument("--folder", default=".")
    parser.add_argument("--delay-sheet", default=None)
    args = parser.parse_args()
    pack_phyto_allInfo_light(
        args.stream_info_h5,
        args.params_h5,
        args.tag_delay_file,
        args.data_name,
        args.folder,
        args.delay_sheet,
    )
